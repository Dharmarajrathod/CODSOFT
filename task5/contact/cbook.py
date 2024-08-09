import tkinter as tk
from tkinter import messagebox, simpledialog
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook # type: ignore

class ContactApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Contact Manager")
        self.root.geometry("375x667")  # Mobile-sized window

        self.contacts = []

        # Main Frame
        self.main_frame = tk.Frame(root, padx=10, pady=10)
        self.main_frame.pack(expand=True, fill=tk.BOTH)

        # Add Contact Frame
        self.add_frame = tk.LabelFrame(self.main_frame, text="Add Contact", padx=10, pady=10)
        self.add_frame.pack(padx=10, pady=10, fill=tk.X)

        tk.Label(self.add_frame, text="Name:").grid(row=0, column=0, sticky=tk.W)
        self.name_entry = tk.Entry(self.add_frame)
        self.name_entry.grid(row=0, column=1, pady=5, sticky=tk.EW)

        tk.Label(self.add_frame, text="Phone Number:").grid(row=1, column=0, sticky=tk.W)
        self.phone_entry = tk.Entry(self.add_frame)
        self.phone_entry.grid(row=1, column=1, pady=5, sticky=tk.EW)

        tk.Label(self.add_frame, text="Email:").grid(row=2, column=0, sticky=tk.W)
        self.email_entry = tk.Entry(self.add_frame)
        self.email_entry.grid(row=2, column=1, pady=5, sticky=tk.EW)

        tk.Label(self.add_frame, text="Address:").grid(row=3, column=0, sticky=tk.W)
        self.address_entry = tk.Entry(self.add_frame)
        self.address_entry.grid(row=3, column=1, pady=5, sticky=tk.EW)

        tk.Button(self.add_frame, text="Add Contact", command=self.add_contact).grid(row=4, columnspan=2, pady=10)

        # Search Frame
        self.search_frame = tk.LabelFrame(self.main_frame, text="Search", padx=10, pady=10)
        self.search_frame.pack(padx=10, pady=10, fill=tk.X)

        tk.Label(self.search_frame, text="Search:").grid(row=0, column=0, sticky=tk.W)
        self.search_entry = tk.Entry(self.search_frame)
        self.search_entry.grid(row=0, column=1, pady=5, sticky=tk.EW)

        tk.Button(self.search_frame, text="Search", command=self.search_contact).grid(row=0, column=2, padx=5)

        # Contact List Frame
        self.list_frame = tk.Frame(self.main_frame, padx=10, pady=10)
        self.list_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.contact_list = tk.Listbox(self.list_frame)
        self.contact_list.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scrollbar = tk.Scrollbar(self.list_frame, orient=tk.VERTICAL, command=self.contact_list.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.contact_list.config(yscrollcommand=self.scrollbar.set)
        self.contact_list.bind('<Double-1>', self.edit_contact)

        # Actions Frame
        self.actions_frame = tk.Frame(root, padx=10, pady=10)
        self.actions_frame.pack(fill=tk.X)

        tk.Button(self.actions_frame, text="Delete Contact", command=self.delete_contact).pack(side=tk.LEFT, padx=5)
        tk.Button(self.actions_frame, text="Save to XML", command=lambda: self.save_to_xml("contacts.xml")).pack(side=tk.LEFT, padx=5)
        tk.Button(self.actions_frame, text="Load from XML", command=lambda: self.load_from_xml("contacts.xml")).pack(side=tk.LEFT, padx=5)
        tk.Button(self.actions_frame, text="Save to Excel", command=lambda: self.save_to_excel("contacts.xlsx")).pack(side=tk.LEFT, padx=5)
        tk.Button(self.actions_frame, text="Load from Excel", command=lambda: self.load_from_excel("contacts.xlsx")).pack(side=tk.LEFT, padx=5)

    def add_contact(self):
        name = self.name_entry.get()
        phone = self.phone_entry.get()
        email = self.email_entry.get()
        address = self.address_entry.get()
        if name and phone:
            self.contacts.append({'name': name, 'phone': phone, 'email': email, 'address': address})
            self.update_contact_list()
        else:
            messagebox.showerror("Error", "Name and phone number are required")

    def update_contact_list(self):
        self.contact_list.delete(0, tk.END)
        for contact in self.contacts:
            self.contact_list.insert(tk.END, f"{contact['name']} ({contact['phone']})")

    def search_contact(self):
        query = self.search_entry.get().lower()
        results = [f"{contact['name']} ({contact['phone']})" for contact in self.contacts if query in contact['name'].lower() or query in contact['phone']]
        self.contact_list.delete(0, tk.END)
        for result in results:
            self.contact_list.insert(tk.END, result)

    def edit_contact(self, event):
        selected = self.contact_list.curselection()
        if selected:
            index = selected[0]
            contact = self.contacts[index]
            new_name = simpledialog.askstring("Edit Name", "New name:", initialvalue=contact['name'])
            new_phone = simpledialog.askstring("Edit Phone", "New phone number:", initialvalue=contact['phone'])
            if new_name and new_phone:
                self.contacts[index] = {'name': new_name, 'phone': new_phone, 'email': contact['email'], 'address': contact['address']}
                self.update_contact_list()

    def delete_contact(self):
        selected = self.contact_list.curselection()
        if selected:
            index = selected[0]
            self.contacts.pop(index)
            self.update_contact_list()

    def save_to_xml(self, filename):
        root = ET.Element("Contacts")
        for contact in self.contacts:
            contact_elem = ET.SubElement(root, "Contact")
            for key, value in contact.items():
                child = ET.SubElement(contact_elem, key)
                child.text = value
        tree = ET.ElementTree(root)
        tree.write(filename)

    def load_from_xml(self, filename):
        self.contacts = []
        tree = ET.parse(filename)
        root = tree.getroot()
        for contact_elem in root.findall('Contact'):
            contact = {}
            for child in contact_elem:
                contact[child.tag] = child.text
            self.contacts.append(contact)
        self.update_contact_list()

    def save_to_excel(self, filename):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Email", "Address"])
        for contact in self.contacts:
            ws.append([contact.get("name", ""), contact.get("phone", ""), contact.get("email", ""), contact.get("address", "")])
        wb.save(filename)

    def load_from_excel(self, filename):
        self.contacts = []
        wb = load_workbook(filename)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            contact = {
                "name": row[0],
                "phone": row[1],
                "email": row[2],
                "address": row[3]
            }
            self.contacts.append(contact)
        self.update_contact_list()

if __name__ == "__main__":
    root = tk.Tk()
    app = ContactApp(root)
    root.mainloop()
